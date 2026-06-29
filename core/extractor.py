from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.webdriver.common.by import By
from core.selectors import CONTAINER_CONVERSA, REMETENTE, MENSAGEM, HORARIO, NAO_LIDAS
from datetime import datetime
import os
import spacy

try:
    nlp = spacy.load("pt_core_news_sm")
except:
    print(
        "   ⚠️ spaCy model not found. Install with: python -m spacy download pt_core_news_sm"
    )
    nlp = None


def classify_message(text):
    if nlp is None:
        return "Outros"
    doc = nlp(text.lower())
    lemmas = [token.lemma_ for token in doc]

    if any(
        p in lemmas
        for p in ["querer", "pedir", "comprar", "encomendar", "gostar", "precisar"]
    ):
        return "Pedidos"
    elif any(
        p in lemmas
        for p in [
            "problema",
            "defeito",
            "reclamar",
            "erro",
            "ruim",
            "falha",
            "queimar",
            "atrasar",
        ]
    ):
        return "Reclamações"
    elif any(
        p in lemmas
        for p in ["entregar", "motoboy", "retirar", "endereço", "chegar", "sair"]
    ):
        return "Entregas"
    elif any(
        p in lemmas
        for p in ["orçamento", "estoque", "preço", "fornecedor", "lote", "reposição"]
    ):
        return "Fornecedores"
    elif any(p in lemmas for p in ["oi", "olá", "bom", "tudo", "beleza"]):
        return "Saudações"
    else:
        return "Outros"


def extract_containers(driver, filename):
    print(f"   📊 Extracting containers... (file: {filename})")

    if not os.path.exists(filename):
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        abas = [
            "Pedidos",
            "Reclamações",
            "Entregas",
            "Fornecedores",
            "Saudações",
            "Outros",
        ]
        for aba in abas:
            ws = wb.create_sheet(aba)
            headers = ["SENDER", "MESSAGE", "TIME", "UNREAD", "EXTRACTION_TIMESTAMP"]
            for col, title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(
                    start_color="4F81BD", end_color="4F81BD", fill_type="solid"
                )
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 25
    else:
        wb = load_workbook(filename)

    last_messages = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0] and row[1]:
                    last_messages[(sheet_name, row[0])] = row[1]

    extraction_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_added = 0

    try:
        containers = driver.find_elements(By.XPATH, CONTAINER_CONVERSA)
        print(f"   📦 Containers found: {len(containers)}")

        for i, container in enumerate(containers, 1):
            try:
                sender = container.find_element(By.XPATH, REMETENTE).text
            except:
                sender = "UNKNOWN"
            try:
                message = container.find_element(By.XPATH, MENSAGEM).text
            except:
                message = ""
            try:
                time = container.find_element(By.XPATH, HORARIO).text
            except:
                time = ""
            try:
                unread = container.find_element(By.XPATH, NAO_LIDAS).text
            except:
                unread = "0"

            if not message:
                continue

            target_sheet = classify_message(message)
            key = (target_sheet, sender)
            if key in last_messages and last_messages[key] == message:
                print(f"   ⏭️  Container {i}: duplicate message (ignored)")
                continue

            ws_target = wb[target_sheet]
            next_row = ws_target.max_row + 1
            ws_target.cell(row=next_row, column=1, value=sender)
            ws_target.cell(row=next_row, column=2, value=message)
            ws_target.cell(row=next_row, column=3, value=time)
            ws_target.cell(row=next_row, column=4, value=unread)
            ws_target.cell(row=next_row, column=5, value=extraction_timestamp)
            total_added += 1
            last_messages[key] = message
            print(
                f"   ✅ Container {i}: {sender} - {message[:30]}... (NEW in {target_sheet})"
            )

    except Exception as e:
        print(f"   ❌ Error extracting: {e}")

    try:
        wb.save(filename)
        print(f"   ✅ Spreadsheet updated: {filename}")
    except PermissionError:
        print("\n" + "=" * 60)
        print("⚠️  FILE OPEN DETECTED!")
        print(f"📌 Close the file '{filename}' in Excel.")
        print("=" * 60 + "\n")
    except Exception as e:
        print(f"   ❌ Error saving: {e}")

    print(f"   📊 Total added: {total_added}")
    return filename
